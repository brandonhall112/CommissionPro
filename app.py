import sys, math
from PySide6.QtGui import QDesktopServices
from PySide6.QtCore import QUrl
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Dict as TDict


def resolve_excel_path(expected_name: str = "Tech days and quote rates.xlsx") -> Path | None:
    """Find the default Excel workbook inside assets without prompting the user unless missing."""
    assets = resolve_assets_dir()

    exact = (assets / expected_name)
    try:
        if exact.exists():
            return exact.resolve()
    except Exception:
        pass

    # Fuzzy match (handles minor renames like '(1)' etc.)
    try:
        for f in assets.glob("*.xlsx"):
            n = f.name.lower()
            if "tech" in n and "quote" in n and "rate" in n:
                return f.resolve()
    except Exception:
        pass

    # If exactly one xlsx exists, use it
    try:
        xls = list(assets.glob("*.xlsx"))
        if len(xls) == 1:
            return xls[0].resolve()
    except Exception:
        pass

    return None



def resolve_assets_dir() -> Path:
    """Return the assets directory for dev + PyInstaller (onefile/onedir).

    onefile: extracted to sys._MEIPASS/assets
    onedir: typically <exe_dir>/_internal/assets (new PyInstaller layout) or <exe_dir>/assets
    dev:   <repo>/assets
    """
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        return (Path(meipass).resolve() / "assets").resolve()

    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        for p in (exe_dir / "_internal" / "assets", exe_dir / "assets"):
            try:
                if p.exists():
                    return p.resolve()
            except Exception:
                pass
        return (exe_dir / "_internal" / "assets").resolve()

    return (Path(__file__).resolve().parent / "assets").resolve()


import numpy as np
import openpyxl

from PySide6.QtCore import Qt, QSize, QRectF
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QSpinBox,
    QComboBox, QCheckBox, QFrame, QScrollArea, QSplitter,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QSizePolicy,
    QTextBrowser, QDialog, QLineEdit, QTextEdit
)
from PySide6.QtPrintSupport import QPrinter, QPrintPreviewDialog
from PySide6.QtGui import QTextDocument
from PySide6.QtGui import QPageSize, QFont, QColor
import base64

APP_TITLE = "Pearson Commissioning Pro"

# Business rules
TRAINING_MACHINES_PER_DAY = 3  # 1 training day per 3 machines (ceil)
DEFAULT_INSTALL_WINDOW = 7
MIN_INSTALL_WINDOW = 3
MAX_INSTALL_WINDOW = 14
TRAVEL_DAYS_PER_PERSON = 2  # travel-in + travel-out
WORKLOAD_CALENDAR_DAYS = 14  # 2-week calendar horizon (Sun-Sat)

# Requested overrides
OVERRIDE_AIRFARE_PER_PERSON = 1500.0
OVERRIDE_BAGGAGE_PER_DAY_PER_PERSON = 150.0

ASSETS_DIR = Path(__file__).resolve().parent / "assets"
DEFAULT_EXCEL = ASSETS_DIR / "Tech days and quote rates.xlsx"
SKILLS_MATRIX_EXCEL = ASSETS_DIR / "Machine Qualifications for PCP Quoting.xlsx"
LOGO_PATH = ASSETS_DIR / "Pearson Logo.png"

RPC_MODELS = {"RPC-C", "RPC-DF", "RPC-PH", "RPC-OU"}
RPC_ENGINEER_TUESDAY_MODELS = {"RPC-PH", "RPC-OU"}
GENERIC_MODEL_ALIASES = {
    "CONV",
    "PRODUCTION SUPPORT DAY",
    "TRAINING DAY",
}


def ceil_int(x: float) -> int:
    return int(math.ceil(float(x)))


def balanced_allocate(total_days: int, headcount: int) -> List[int]:
    """Balance integer days to minimize the maximum assigned days."""
    if headcount <= 0:
        return []
    loads = [0] * headcount
    for _ in range(int(total_days)):
        i = int(np.argmin(loads))
        loads[i] += 1
    loads.sort(reverse=True)
    return loads




def chunk_allocate_by_machine(install_days_per_machine: int, qty: int, training_days: int, window: int) -> List[int]:
    """Allocate work using whole-machine install chunks + whole-day training chunks.

    Install days are assigned per machine (no fractional splitting). Training days are 1-day chunks.
    We choose the *minimum* headcount that keeps every person's onsite days <= window.

    Training assignment heuristic:
      - Prefer assigning training to the currently *most-loaded* person that can still accept a day
        without exceeding the window (keeps extra people from traveling and mirrors reality).
      - If none can accept, fall back to the least-loaded person (best-effort).
    Returns a list of total onsite days per person (sorted descending).
    """
    install_days_per_machine = int(install_days_per_machine or 0)
    qty = int(qty or 0)
    training_days = int(training_days or 0)
    window = int(window or 0)

    if window <= 0:
        return []
    if qty <= 0 and training_days <= 0:
        return []

    # If there is no install work, allocate training only.
    if qty <= 0 or install_days_per_machine <= 0:
        headcount = ceil_int(training_days / window) if training_days > 0 else 0
        loads = balanced_allocate(training_days, headcount) if headcount > 0 else []
        return loads

    max_headcount = max(1, qty)  # at most one machine per person
    for headcount in range(1, max_headcount + 1):
        # Distribute whole machines as evenly as possible.
        base_n = qty // headcount
        rem = qty % headcount
        machine_counts = [base_n + (1 if i < rem else 0) for i in range(headcount)]
        loads = [c * install_days_per_machine for c in machine_counts]

        # Assign training days as 1-day chunks.
        for _ in range(training_days):
            # Prefer adding to the most-loaded person who can still accept 1 day within window.
            candidates = [i for i, d in enumerate(loads) if d + 1 <= window]
            if candidates:
                i = max(candidates, key=lambda j: loads[j])
            else:
                i = int(np.argmin(loads))
            loads[i] += 1

        if max(loads) <= window:
            loads.sort(reverse=True)
            return loads

    # Best-effort fallback (should generally be prevented by validation).
    loads = [install_days_per_machine] * qty
    for _ in range(training_days):
        i = int(np.argmin(loads))
        loads[i] += 1
    loads.sort(reverse=True)
    return loads

@dataclass
class ModelInfo:
    item: str
    tech_install_days_per_machine: int   # install-only (no training baked in)
    eng_days_per_machine: int
    training_applicable: bool = True


@dataclass
class LineSelection:
    model: str
    qty: int
    training_required: bool


@dataclass
class RoleTotals:
    headcount: int
    total_onsite_days: int
    onsite_days_by_person: List[int]
    day_rate: float
    labor_cost: float


@dataclass
class ExpenseLine:
    description: str
    quantity: float
    unit_price: float
    extended: float
    details: str


@dataclass
class Assignment:
    model: str
    role: str  # "Technician" or "Engineer"
    person_num: int
    onsite_days: int
    cost: float
    crew_pool: str = ""


class SkillsMatrix:
    """Technician qualification matrix loader used for tech-only grouping."""

    def __init__(self, path: Path):
        self.path = path
        self.tech_rows: List[Dict[str, str]] = []
        self.model_headers: set[str] = set()
        self._load()

    def _load(self):
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = wb.active

        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            headers.append(str(v).strip() if v is not None else "")

        model_cols: Dict[int, str] = {}
        for idx, name in enumerate(headers, 1):
            if idx <= 2:
                continue
            if name:
                model_cols[idx] = name
                self.model_headers.add(name)

        for r in range(2, ws.max_row + 1):
            resource_type = ws.cell(r, 1).value
            if str(resource_type).strip().lower() != "technician":
                continue

            ratings: Dict[str, str] = {}
            for c, model in model_cols.items():
                val = ws.cell(r, c).value
                ratings[model] = str(val).strip().upper() if val is not None else ""
            self.tech_rows.append(ratings)

    def has_model(self, model: str) -> bool:
        return model in self.model_headers

    def can_group_models(self, models: List[str]) -> bool:
        """Rule: >=2 techs with T3 on all models, and >=1 tech with >=T2 on all models."""
        if not models:
            return True

        t3_count = 0
        t2_plus_count = 0
        for row in self.tech_rows:
            vals = [row.get(m, "") for m in models]
            if all(v == "T3" for v in vals):
                t3_count += 1
            if all(v in ("T2", "T3") for v in vals):
                t2_plus_count += 1
        return t3_count >= 2 and t2_plus_count >= 1


class ExcelData:
    def __init__(self, path: Path):
        self.path = path
        self.models: Dict[str, ModelInfo] = {}
        self.rates: Dict[str, Dict[str, object]] = {}
        self.requirements: List[str] = []
        self._load()

    def _load(self):
        wb = openpyxl.load_workbook(self.path, data_only=True)

        # Models: Instal days by Model
        if "Instal days by Model" not in wb.sheetnames:
            raise ValueError("Missing sheet: 'Instal days by Model'")
        ws = wb["Instal days by Model"]

        headers = {
            str(ws.cell(1, c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value is not None
        }

        def find_col(pred):
            for k, c in headers.items():
                if pred(k.lower()):
                    return c
            return None

        col_item = find_col(lambda s: s in ["item", "model", "machine", "machine type"])
        col_tech = find_col(lambda s: "technician" in s and "day" in s)
        col_eng = find_col(lambda s: ("engineer" in s and "day" in s) or ("field engineer" in s and "day" in s))
        col_train_app = find_col(lambda s: ("training required" in s))

        if col_item is None or col_tech is None or col_eng is None:
            raise ValueError("Model sheet columns not found. Expected: Item, Technician Days Required, Field Engineer Days Required.")

        def _as_bool(v, default=True):
            if v is None:
                return default
            if isinstance(v, bool):
                return v
            s = str(v).strip().lower()
            if s in ("true","t","yes","y","1"):
                return True
            if s in ("false","f","no","n","0"):
                return False
            return default

        for r in range(2, ws.max_row + 1):
            item = ws.cell(r, col_item).value
            if item is None:
                continue
            item = str(item).strip()
            if not item:
                continue
            tech = ws.cell(r, col_tech).value or 0
            eng = ws.cell(r, col_eng).value or 0
            try:
                tech_i = int(float(tech))
            except Exception:
                tech_i = 0
            try:
                eng_i = int(float(eng))
            except Exception:
                eng_i = 0
            train_app = _as_bool(ws.cell(r, col_train_app).value, default=True) if col_train_app is not None else True
            self.models[item] = ModelInfo(item=item, tech_install_days_per_machine=tech_i, eng_days_per_machine=eng_i, training_applicable=train_app)

        # Rates: Service Rates
        if "Service Rates" not in wb.sheetnames:
            raise ValueError("Missing sheet: 'Service Rates'")
        ws = wb["Service Rates"]

        header_row = None
        for r in range(1, 15):
            if str(ws.cell(r, 2).value).strip().lower() == "item" and str(ws.cell(r, 3).value).strip().lower() == "description":
                header_row = r
                break
        if header_row is None:
            header_row = 3

        for r in range(header_row + 1, ws.max_row + 1):
            desc = ws.cell(r, 3).value
            if desc is None:
                continue
            desc_s = str(desc).strip()
            if not desc_s:
                continue
            unit = ws.cell(r, 6).value
            notes = ws.cell(r, 7).value
            try:
                unit_f = float(unit)
            except Exception:
                continue
            self.rates[desc_s.lower()] = {
                "description": desc_s,
                "unit_price": unit_f,
                "notes": str(notes).strip() if notes is not None else ""
            }

        # Requirements & Assumptions
        if "Requirements and Assumptions" in wb.sheetnames:
            ws = wb["Requirements and Assumptions"]
            out = []
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 3).value
                if v is None:
                    continue
                s = str(v).strip()
                if s and not s.lower().startswith("assumptions and requirements"):
                    out.append(s)
            self.requirements = out

    def get_rate(self, key: str) -> Tuple[float, str]:
        k = key.lower().strip()
        if k in self.rates:
            rv = self.rates[k]
            return float(rv["unit_price"]), str(rv["description"])
        for rk, rv in self.rates.items():
            if k in rk:
                return float(rv["unit_price"]), str(rv["description"])
        raise KeyError(f"Rate not found for '{key}'")


class MachineLine(QFrame):
    def __init__(self, models: List[str], training_applicable_map: Dict[str, bool], on_change, on_delete):
        super().__init__()
        self.on_change = on_change
        self.on_delete = on_delete
        self.training_applicable_map = training_applicable_map

        self.setObjectName("machineLine")
        row = QHBoxLayout(self)
        row.setContentsMargins(10, 10, 10, 10)
        row.setSpacing(10)

        self.cmb_model = QComboBox()
        self.cmb_model.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.cmb_model.addItem("— Select —")
        self.cmb_model.addItems(models)
        self.cmb_model.setCurrentIndex(-1)
        self.cmb_model.currentIndexChanged.connect(self._model_changed)

        self.spin_qty = QSpinBox()
        self.spin_qty.setRange(0, 999)
        self.spin_qty.setValue(1)
        self.spin_qty.valueChanged.connect(self._changed)

        self.chk_training = QCheckBox("Training Required")
        self.chk_training.setChecked(True)
        self.chk_training.stateChanged.connect(self._changed)
        self.chk_training.setToolTip("Uncheck only by customer request (training is normally included).")

        self.btn_delete = QPushButton("🗑")
        self.btn_delete.setFixedWidth(40)
        self.btn_delete.clicked.connect(self._delete)

        row.addWidget(QLabel("Machine Model"))
        row.addWidget(self.cmb_model, 2)
        row.addWidget(QLabel("Qty"))
        row.addWidget(self.spin_qty)
        row.addWidget(self.chk_training, 1)
        row.addWidget(self.btn_delete)

        self._model_changed()

    def _set_training_visibility(self, model: str):
        """Show/hide training checkbox without mutating checked state."""
        if model == "— Select —":
            model = ""
        model = model.strip()
        if not model:
            self.chk_training.hide()
            return

        applicable = bool(self.training_applicable_map.get(model, True))
        if not applicable:
            self.chk_training.hide()
            return

        self.chk_training.show()

    def _model_changed(self, *_):
        self._set_training_visibility(self.cmb_model.currentText())
        self.on_change()

    def _changed(self, *_):
        self.on_change()

    def _delete(self):
        self.on_delete(self)

    def value(self) -> LineSelection:
        model = self.cmb_model.currentText().strip()
        if model == "— Select —":
            model = ""
        return LineSelection(
            model=model,
            qty=int(self.spin_qty.value()) if model else 0,
            training_required=(bool(self.chk_training.isChecked()) if self.chk_training.isVisible() else False)
        )



class Card(QFrame):
    def __init__(self, title: str, icon_text: str):
        super().__init__()
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setObjectName("card")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 12)

        top = QHBoxLayout()
        ic = QLabel(icon_text)
        ic.setObjectName("cardIcon")
        ic.setFixedSize(QSize(34, 34))
        ic.setAlignment(Qt.AlignCenter)
        top.addWidget(ic)

        v = QVBoxLayout()
        self.lbl_title = QLabel(title)
        self.lbl_title.setObjectName("cardTitle")
        self.lbl_value = QLabel("—")
        self.lbl_value.setObjectName("cardValue")
        self.lbl_sub = QLabel("")
        self.lbl_sub.setObjectName("cardSub")
        self.lbl_sub.setWordWrap(True)
        v.addWidget(self.lbl_title)
        v.addWidget(self.lbl_value)
        top.addLayout(v, 1)

        lay.addLayout(top)
        lay.addWidget(self.lbl_sub)

    def set_value(self, value: str, sub: str = ""):
        self.lbl_value.setText(value)
        self.lbl_sub.setText(sub or "")


class Section(QFrame):
    def __init__(self, title: str, subtitle: str = "", icon_text: str = ""):
        super().__init__()
        self.setObjectName("section")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 14, 14, 14)
        lay.setSpacing(10)

        head = QHBoxLayout()
        if icon_text:
            ic = QLabel(icon_text)
            ic.setObjectName("sectionIcon")
            ic.setFixedSize(QSize(28, 28))
            ic.setAlignment(Qt.AlignCenter)
            head.addWidget(ic)

        title_box = QVBoxLayout()
        t = QLabel(title)
        t.setObjectName("sectionTitle")
        title_box.addWidget(t)
        if subtitle:
            s = QLabel(subtitle)
            s.setObjectName("sectionSub")
            s.setWordWrap(True)
            title_box.addWidget(s)
        head.addLayout(title_box, 1)
        lay.addLayout(head)

        self.content = QWidget()
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(10)
        lay.addWidget(self.content)


def money(x: float) -> str:
    return f"${x:,.0f}"


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1920, 1200)

        self.data = ExcelData(DEFAULT_EXCEL)
        self.models_sorted = sorted(self.data.models.keys())
        self.training_app_map = {k: bool(v.training_applicable) for k, v in self.data.models.items()}
        self.lines: List[MachineLine] = []
        self.quote_header_fields = {
            "customer_name": "",
            "reference": "",
            "submitted_to": "",
            "prepared_by": "",
        }

        self.skills_matrix: SkillsMatrix | None = None
        self.skills_warning = ""
        self._load_skills_matrix()

        central_container = QWidget()
        root = QVBoxLayout(central_container)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Outer scroll area enables whole-window scrolling in stacked (single-column) mode
        self.outer_scroll = QScrollArea()
        self.outer_scroll.setObjectName("outerScroll")
        self.outer_scroll.setWidgetResizable(True)
        self.outer_scroll.setFrameShape(QFrame.NoFrame)
        self.outer_scroll.setWidget(central_container)
        self.setCentralWidget(self.outer_scroll)


        header = QFrame()
        header.setObjectName("header")
        h = QHBoxLayout(header)
        h.setContentsMargins(14, 10, 14, 10)
        self.lbl_title = QLabel("🧾  " + APP_TITLE)
        self.lbl_title.setObjectName("appTitle")
        h.addWidget(self.lbl_title)
        h.addStretch(1)
        btn_header = QPushButton("Header")
        btn_header.setToolTip("Enter quote header details (customer/reference/submitted to/prepared by).")
        btn_header.clicked.connect(self.open_header_form)
        h.addWidget(btn_header)

        btn_load = QPushButton("Load Excel…")
        btn_load.setToolTip("Load a different Excel workbook (will replace the bundled rates/models for this session).")
        btn_load.clicked.connect(self.open_excel)
        h.addWidget(btn_load)

        btn_open_bundled = QPushButton("Open Bundled Excel")
        btn_open_bundled.setToolTip("Open the Excel workbook that was bundled into this EXE (for verification).")
        btn_open_bundled.clicked.connect(self.open_bundled_excel)
        h.addWidget(btn_open_bundled)

        btn_help = QPushButton("Help")
        btn_help.setToolTip("Open the user guide and calculation notes (README).")
        btn_help.clicked.connect(self.open_help)
        h.addWidget(btn_help)
        root.addWidget(header)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        self.splitter = splitter
        root.addWidget(splitter, 1)

        # LEFT
        left = QFrame()
        left.setObjectName("panel")
        left_l = QVBoxLayout(left)
        left_l.setContentsMargins(14, 14, 14, 14)
        left_l.setSpacing(12)

        t = QLabel("Machine Configuration")
        t.setObjectName("panelTitle")
        left_l.addWidget(t)

        left_l.addWidget(QLabel(
            "Add machines to estimate commissioning requirements.\\n"
            "Each machine type requires dedicated personnel (no sharing across types)."
        ))

        win_box = QFrame()
        win_box.setObjectName("softBox")
        win_l = QHBoxLayout(win_box)
        win_l.setContentsMargins(12, 10, 12, 10)
        win_l.addWidget(QLabel("Customer Install Window"))
        self.spin_window = QSpinBox()
        self.spin_window.setRange(MIN_INSTALL_WINDOW, MAX_INSTALL_WINDOW)
        self.spin_window.setValue(DEFAULT_INSTALL_WINDOW)
        self.spin_window.valueChanged.connect(self.recalc)
        win_l.addStretch(1)
        win_l.addWidget(self.spin_window)
        win_l.addWidget(QLabel("days"))
        left_l.addWidget(win_box)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        container = QWidget()
        self.lines_layout = QVBoxLayout(container)
        self.lines_layout.setContentsMargins(0, 0, 0, 0)
        self.lines_layout.setSpacing(10)

        self.empty_hint = QLabel("No machines added.\\nClick “Add Machine” to begin.")
        self.empty_hint.setObjectName("emptyHint")
        self.empty_hint.setAlignment(Qt.AlignCenter)
        self.empty_hint.setMinimumHeight(120)
        self.lines_layout.addWidget(self.empty_hint)

        self.scroll.setWidget(container)
        self.scroll.setMinimumHeight(240)
        left_l.addWidget(self.scroll, 1)

        btn_add = QPushButton("+  Add Machine")
        btn_add.setObjectName("addMachine")
        btn_add.clicked.connect(self.add_line)
        left_l.addWidget(btn_add)

        note = QLabel("Note: Unchecking “Training Required” should only be done by customer request.")
        note.setObjectName("note")
        note.setWordWrap(True)
        left_l.addWidget(note)

        splitter.addWidget(left)

        # RIGHT (scrollable)
        right_wrap = QWidget()
        self.right_wrap = right_wrap
        right_layout = QVBoxLayout(right_wrap)
        right_layout.setContentsMargins(0, 0, 0, 0)

        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        self.right_scroll = right_scroll
        right_layout.addWidget(right_scroll)

        right = QWidget()
        self.right_content = right
        right_scroll.setWidget(right)
        right_l = QVBoxLayout(right)
        right_l.setContentsMargins(14, 14, 14, 14)
        right_l.setSpacing(12)

        cards = QHBoxLayout()
        self.card_tech = Card("Technicians", "🧰")
        self.card_eng = Card("Engineers", "🧑‍💻")
        self.card_window = Card("Max Onsite", "⏱")
        self.card_total = Card("Total Cost", "💲")
        cards.addWidget(self.card_tech, 1)
        cards.addWidget(self.card_eng, 1)
        cards.addWidget(self.card_window, 1)
        cards.addWidget(self.card_total, 1)
        right_l.addLayout(cards)

        self.alert = QLabel("")
        self.alert.setObjectName("alert")
        self.alert.setWordWrap(True)
        self.alert.hide()
        right_l.addWidget(self.alert)

        self.tbl_breakdown = self.make_table(["Model", "Qty", "Tech Days", "Eng Days", "Technicians", "Engineers"])
        self.tbl_assign = self.make_table(["Machine Type", "Role", "Person #", "Assigned Days", "Cost"])
        self.tbl_labor = self.make_table(["Role", "Daily Rate", "Total Days", "Personnel", "Total Cost"])
        self.tbl_exp = self.make_table(["Expense", "Details", "Amount"])
        self.tbl_exp.setMinimumHeight(0)

        sec_breakdown = Section("Machine Breakdown", "Days and personnel required per machine model", "🧩")
        sec_breakdown.content_layout.addWidget(self.tbl_breakdown)

        sec_assign = Section("Personnel Assignments", "Each machine type has dedicated personnel.", "👥")
        sec_assign.content_layout.addWidget(self.tbl_assign)

        sec_labor = Section("Labor Costs", "Labor costs by role at daily rates (8 hours/day).", "🛠")
        sec_labor.content_layout.addWidget(self.tbl_labor)

        # Workload calendar in Gantt style (2-week Sun-Sat view)
        self.tbl_workload_calendar = QTableWidget(6, WORKLOAD_CALENDAR_DAYS)
        self.tbl_workload_calendar.setObjectName("table")
        self.tbl_workload_calendar.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_workload_calendar.setSelectionMode(QAbstractItemView.NoSelection)
        self.tbl_workload_calendar.horizontalHeader().setDefaultSectionSize(24)
        self.tbl_workload_calendar.verticalHeader().setDefaultSectionSize(26)
        # Stretch day columns to consume available width now that the horizon is 14 days.
        self.tbl_workload_calendar.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        cal_header_font = self.tbl_workload_calendar.horizontalHeader().font()
        cal_header_font.setPointSizeF(max(7.0, cal_header_font.pointSizeF() - 1.0))
        self.tbl_workload_calendar.horizontalHeader().setFont(cal_header_font)
        self.tbl_workload_calendar.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.tbl_workload_calendar.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.tbl_workload_calendar.setHorizontalHeaderLabels([
            ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"][i % 7] for i in range(WORKLOAD_CALENDAR_DAYS)
        ])
        self.tbl_workload_calendar.setMinimumHeight(320)
        self.tbl_workload_calendar.setToolTip(
            "Gantt trip view (2 weeks): light color = travel day, solid color = onsite day. "
            "Tech uses #e04426 and Engineer uses #6790a0."
        )
        sec_chart = Section("Workload Calendar", "2-week Sun-Sat Gantt trip view.", "🗓️")
        sec_chart.content_layout.addWidget(self.tbl_workload_calendar)
        cal_legend = QLabel("Legend: Light shade = travel day · Dark shade = onsite day")
        cal_legend.setObjectName("muted")
        sec_chart.content_layout.addWidget(cal_legend)

        # Left side: put calendar under Machine Configuration so the right-side widgets stay readable
        left_l.addWidget(sec_chart)

        right_l.addWidget(sec_breakdown)
        right_l.addWidget(sec_assign)
        right_l.addWidget(sec_labor)

        sec_exp = Section("Estimated Expenses", "", "🧳")
        self.lbl_exp_hdr = QLabel("")
        self.lbl_exp_hdr.setObjectName("sectionSub")
        self.lbl_exp_hdr.setWordWrap(True)
        sec_exp.content_layout.addWidget(self.lbl_exp_hdr)
        sec_exp.content_layout.addWidget(self.tbl_exp)
        right_l.addWidget(sec_exp)

        bottom = QFrame()
        bottom.setObjectName("totalBar")
        bl = QHBoxLayout(bottom)
        bl.setContentsMargins(14, 12, 14, 12)
        self.lbl_total = QLabel("Estimated Total")
        self.lbl_total.setObjectName("totalLabel")
        self.lbl_total_val = QLabel("—")
        self.lbl_total_val.setObjectName("totalValue")
        bl.addWidget(self.lbl_total)
        bl.addStretch(1)
        bl.addWidget(self.lbl_total_val)
        self.btn_print = QPushButton("Print Quote…")
        self.btn_print.clicked.connect(self.print_quote_preview)
        self.btn_print.setEnabled(False)
        bl.addWidget(self.btn_print)
        right_l.addWidget(bottom)

        splitter.addWidget(right_wrap)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)

        self.apply_theme()
        self.reset_views()

        # Responsive scaling baseline (designed for 1920x1200)
        self._base_font_pt = float(self.font().pointSizeF() or 10.0)
        self._apply_scale()

        # Responsive layout: two-column w/ right scroll on large screens; single stacked w/ full-window scroll on small screens
        self._stack_threshold = 1280
        self._is_stacked = False
        self._apply_responsive_layout()

    def make_table(self, headers: List[str]) -> QTableWidget:
        tbl = QTableWidget(0, len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        tbl.setSelectionMode(QAbstractItemView.SingleSelection)
        tbl.setAlternatingRowColors(True)
        tbl.setObjectName("table")
        tbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        tbl.setMinimumHeight(120)
        return tbl

    def apply_theme(self):
        # Pearson-ish palette (navy + orange + neutral)
        blue = "#4B4F54"   # charcoal gray (logo text)
        gold = "#F05A28"   # Pearson orange
        neutral = "#6D6E71"
        red = "#D6453D"    # Pearson red accent
        css = """
        QFrame#header { background: __BLUE__; color: white; border: none; }
        QLabel#appTitle { color: white; font-size: 20px; font-weight: 800; }
        QFrame#panel { background: white; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#panelTitle { font-size: 16px; font-weight: 800; color: #0F172A; }
        QFrame#softBox { background: #FFF7EA; border: 1px solid #F0D8A8; border-radius: 12px; }
        QLabel#note { color: #334155; font-size: 12px; }
        QLabel#emptyHint { color: __NEUTRAL__; background: #F8FAFC; border: 1px dashed #CBD5E1; border-radius: 12px; }
        QPushButton#primary {
            background: __GOLD__; border: 0px; color: #0B1B2A;
            padding: 10px 12px; border-radius: 10px; font-weight: 800;
        }
        QPushButton#addMachine {
            background: #bebebe; border: 0px; color: #0B1B2A;
            padding: 10px 12px; border-radius: 10px; font-weight: 800;
        }
        QPushButton#addMachine:hover { background: #D6D9DD; }
        QPushButton#addMachine:pressed { background: #CBD5E1; }
        QPushButton {
            padding: 8px 10px; border-radius: 10px;
            border: 1px solid #D6D9DD; background: #F8FAFC;
        }
        QPushButton:disabled { color: #94A3B8; background: #F1F5F9; }
        QFrame#card { background: #FFFFFF; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#cardIcon { background: #EEF2F7; border-radius: 10px; font-size: 16px; }
        QLabel#cardTitle { font-size: 12px; color: __NEUTRAL__; font-weight: 700; }
        QLabel#cardValue { font-size: 24px; color: #0F172A; font-weight: 900; }
        QLabel#cardSub { font-size: 12px; color: __NEUTRAL__; }
        QFrame#section { background: #FFFFFF; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#sectionIcon { background: #EEF2F7; border-radius: 10px; font-size: 14px; }
        QLabel#sectionTitle { font-size: 15px; font-weight: 900; color: #0F172A; }
        QLabel#sectionSub { font-size: 12px; color: __NEUTRAL__; }
        QTableWidget#table {
            background: #FFFFFF;
            border: 1px solid #EEF0F2;
            border-radius: 12px;
            gridline-color: #E2E8F0;
            selection-background-color: #DBEAFE;
        }
        QHeaderView::section {
            background: #343551;
            color: white;
            padding: 8px;
            border: 0px;
            border-bottom: 1px solid #E2E8F0;
            font-weight: 800;
        }
        QFrame#totalBar { background: #F8FAFC; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#totalLabel { font-size: 13px; font-weight: 800; color: #0F172A; }
        QLabel#totalValue { font-size: 22px; font-weight: 900; color: __BLUE__; }
        QLabel#alert {
            background: #FEF2F2;
            border: 1px solid #FCA5A5;
            color: #7F1D1D;
            padding: 10px;
            border-radius: 12px;
        }
        QFrame#machineLine { background: #FFFFFF; border: 1px solid #E6E8EB; border-radius: 12px; }
        """

        css = css.replace("__BLUE__", blue).replace("__GOLD__", gold).replace("__NEUTRAL__", neutral).replace("__RED__", red)
        self.setStyleSheet(css)

    def reset_views(self):
        self.card_tech.set_value("0", "0 total days")
        self.card_eng.set_value("0", "0 total days")
        self.card_window.set_value(f"{self.spin_window.value()}", "")
        self.card_total.set_value("—", "labor + expenses")
        self.lbl_total_val.setText("—")
        for tbl in [self.tbl_breakdown, self.tbl_assign, self.tbl_labor, self.tbl_exp]:
            tbl.setRowCount(0)
        self.lbl_exp_hdr.setText("")
        self.btn_print.setEnabled(False)
        self.alert.hide()
        self.alert.setText("")
        self._clear_workload_calendar()

    def add_line(self):
        if self.empty_hint is not None:
            self.empty_hint.hide()
        ln = MachineLine(self.models_sorted, self.training_app_map, on_change=self.recalc, on_delete=self.delete_line)
        ln.cmb_model.currentIndexChanged.connect(self._refresh_model_choices)
        self.lines.append(ln)
        self.lines_layout.addWidget(ln)
        self._refresh_model_choices()
        self.recalc()

    def delete_line(self, ln: MachineLine):
        self.lines.remove(ln)
        ln.setParent(None)
        ln.deleteLater()
        self._refresh_model_choices()
        if len(self.lines) == 0:
            self.empty_hint.show()
            self.reset_views()
        else:
            self.recalc()

    def _refresh_model_choices(self):
        """Prevent selecting the same machine model on multiple lines."""
        if not self.lines:
            return

        selected = []
        for ln in self.lines:
            v = ln.value().model
            if v:
                selected.append(v)

        for ln in self.lines:
            current = ln.value().model
            ln.cmb_model.blockSignals(True)
            ln.cmb_model.clear()
            ln.cmb_model.addItem("— Select —")
            ln.cmb_model.addItems(self.models_sorted)

            if current and current in self.models_sorted:
                ln.cmb_model.setCurrentIndex(self.models_sorted.index(current) + 1)
            else:
                ln.cmb_model.setCurrentIndex(0)

            # Disable models already chosen on other lines.
            for idx, model in enumerate(self.models_sorted, start=1):
                item = ln.cmb_model.model().item(idx)
                if item is not None:
                    item.setEnabled(not (model in selected and model != current))

            ln.cmb_model.blockSignals(False)

            model = ln.cmb_model.currentText().strip()
            ln.chk_training.blockSignals(True)
            try:
                ln._set_training_visibility(model)
            finally:
                ln.chk_training.blockSignals(False)

    def _load_skills_matrix(self):
        self.skills_matrix = None
        self.skills_warning = ""
        try:
            if SKILLS_MATRIX_EXCEL.exists():
                self.skills_matrix = SkillsMatrix(SKILLS_MATRIX_EXCEL)
            else:
                self.skills_warning = (
                    "Skills matrix file not found; using standard allocation behavior. "
                    "Expected: assets/Machine Qualifications for PCP Quoting.xlsx"
                )
        except Exception as e:
            self.skills_warning = f"Skills matrix unavailable ({e}); using standard allocation behavior."

    @staticmethod
    def _is_generic_model_name(model: str) -> bool:
        upper = model.strip().upper()
        if upper in GENERIC_MODEL_ALIASES:
            return True
        return any(tok in upper for tok in ("CONV", "PRODUCTION SUPPORT", "TRAINING DAY"))

    def _partition_tech_groups(self, selections: List[LineSelection]) -> Dict[str, List[LineSelection]]:
        """Partition tech-only lines into crew pools driven by RPC + skills-matrix rules."""
        groups: Dict[str, List[LineSelection]] = {}

        if not selections:
            return groups

        robot_lines = [s for s in selections if s.model in RPC_MODELS]
        if robot_lines:
            groups["RPC"] = robot_lines

        tech_only = []
        for s in selections:
            mi = self.data.models[s.model]
            if mi.eng_days_per_machine > 0 or s.model in RPC_MODELS:
                continue
            tech_only.append(s)

        generic_lines = [s for s in tech_only if self._is_generic_model_name(s.model)]
        candidate = [s for s in tech_only if s not in generic_lines]

        if self.skills_matrix is None:
            if candidate or generic_lines:
                groups["Tech"] = candidate + generic_lines
            return groups

        skills_known = [s for s in candidate if self.skills_matrix.has_model(s.model)]
        matrix_missing = [s for s in candidate if not self.skills_matrix.has_model(s.model)]

        tech_groups: List[List[LineSelection]] = []
        for line in skills_known:
            placed = False
            for grp in tech_groups:
                models = sorted({x.model for x in grp + [line]})
                if self.skills_matrix.can_group_models(models):
                    grp.append(line)
                    placed = True
                    break
            if not placed:
                tech_groups.append([line])

        supplemental = matrix_missing + generic_lines
        if not tech_groups and supplemental:
            tech_groups = [[]]

        if tech_groups:
            for i, line in enumerate(supplemental):
                tech_groups[i % len(tech_groups)].append(line)
            for i, grp in enumerate(tech_groups, 1):
                groups[f"Tech {i}"] = grp
        elif supplemental:
            groups["Tech 1"] = supplemental

        return groups

    def open_bundled_excel(self):
        """Open the bundled Excel workbook in the user's default spreadsheet app."""
        try:
            p = resolve_excel_path()
            if not p or not p.exists():
                QMessageBox.warning(self, "Bundled Excel not found",
                                    "The bundled Excel file could not be found inside the app assets.")
                return
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(p)))
        except Exception as e:
            QMessageBox.critical(self, "Open error", str(e))


    def open_excel(self):
        fp, _ = QFileDialog.getOpenFileName(self, "Select Excel file", "", "Excel (*.xlsx)")
        if not fp:
            return
        try:
            self.data = ExcelData(Path(fp))
            self.models_sorted = sorted(self.data.models.keys())
            self.training_app_map = {k: bool(v.training_applicable) for k, v in self.data.models.items()}
            for ln in self.lines:
                ln.training_applicable_map = self.training_app_map
            self._refresh_model_choices()
            self.recalc()
        except Exception as e:
            QMessageBox.critical(self, "Excel load error", str(e))

    def _find_readme_path(self) -> Path | None:
        candidates = [
            Path(__file__).resolve().parent / "README.md",
            Path(__file__).resolve().parent.parent / "README.md",
            Path.cwd() / "README.md",
        ]
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidates.extend([
                Path(meipass) / "README.md",
                Path(meipass).parent / "README.md",
            ])
        if getattr(sys, "frozen", False):
            exe_dir = Path(sys.executable).resolve().parent
            candidates.extend([
                exe_dir / "README.md",
                exe_dir / "_internal" / "README.md",
            ])
        for p in candidates:
            try:
                if p.exists():
                    return p
            except Exception:
                continue
        return None

    def open_header_form(self):
        """Collect quote-header details used in printable quote output."""
        dlg = QDialog(self)
        dlg.setWindowTitle("Quote Header")
        dlg.resize(620, 480)
        lay = QVBoxLayout(dlg)

        lay.addWidget(QLabel("Customer Name"))
        customer_in = QLineEdit(dlg)
        customer_in.setPlaceholderText("Customer Name")
        customer_in.setText(str(self.quote_header_fields.get("customer_name", "")))
        lay.addWidget(customer_in)

        lay.addWidget(QLabel("Reference"))
        ref_in = QLineEdit(dlg)
        ref_in.setPlaceholderText("Reference / PO / Project #")
        ref_in.setText(str(self.quote_header_fields.get("reference", "")))
        lay.addWidget(ref_in)

        lay.addWidget(QLabel("Submitted to (name, email, phone)"))
        submitted_in = QTextEdit(dlg)
        submitted_in.setPlaceholderText("Name\nEmail\nPhone")
        submitted_in.setMinimumHeight(120)
        submitted_in.setPlainText(str(self.quote_header_fields.get("submitted_to", "")))
        lay.addWidget(submitted_in)

        lay.addWidget(QLabel("Prepared By"))
        prepared_in = QLineEdit(dlg)
        prepared_in.setPlaceholderText("Prepared by")
        prepared_in.setText(str(self.quote_header_fields.get("prepared_by", "")))
        lay.addWidget(prepared_in)

        actions = QHBoxLayout()
        actions.addStretch(1)
        btn_cancel = QPushButton("Cancel", dlg)
        btn_save = QPushButton("Save", dlg)
        btn_cancel.clicked.connect(dlg.reject)
        btn_save.clicked.connect(dlg.accept)
        actions.addWidget(btn_cancel)
        actions.addWidget(btn_save)
        lay.addLayout(actions)

        if dlg.exec() == QDialog.Accepted:
            self.quote_header_fields["customer_name"] = customer_in.text().strip()
            self.quote_header_fields["reference"] = ref_in.text().strip()
            self.quote_header_fields["submitted_to"] = submitted_in.toPlainText().strip()
            self.quote_header_fields["prepared_by"] = prepared_in.text().strip()

    def open_help(self):
        """Display README guidance in an in-app help dialog."""
        try:
            readme_path = self._find_readme_path()
            if readme_path is not None:
                text = readme_path.read_text(encoding="utf-8", errors="replace")
                title = f"Help — {readme_path.name}"
            else:
                text = (
                    "README.md was not found in this build.\n\n"
                    "This tool estimates commissioning labor, expenses, and total quote values "
                    "from selected machine models and rate data in Excel."
                )
                title = "Help"

            dlg = QDialog(self)
            dlg.setWindowTitle(title)
            dlg.resize(1000, 760)
            lay = QVBoxLayout(dlg)
            viewer = QTextBrowser(dlg)
            viewer.setOpenExternalLinks(True)
            if readme_path is not None:
                try:
                    viewer.setSearchPaths([str(readme_path.parent)])
                except Exception:
                    pass
            viewer.setMarkdown(text)
            lay.addWidget(viewer)
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(self, "Help error", str(e))

    def calc(self):
        selections = [ln.value() for ln in self.lines]
        selections = [s for s in selections if s.qty > 0 and s.model and s.model in self.data.models]
        if not selections:
            raise ValueError("No machines selected. Click “Add Machine” to begin.")

        selected_models = {s.model for s in selections}
        if any(m in RPC_ENGINEER_TUESDAY_MODELS for m in selected_models):
            eng_travel_in_day = 3  # Tuesday
        elif any(m in RPC_MODELS for m in selected_models):
            eng_travel_in_day = 2  # Monday
        else:
            eng_travel_in_day = 1  # Sunday

        window = int(self.spin_window.value())

        tech_hr, _ = self.data.get_rate("tech. regular time")
        eng_hr, _ = self.data.get_rate("eng. regular time")

        def _read_rate_with_fallback(primary_key: str, fallback_key: str) -> float:
            try:
                rate, _ = self.data.get_rate(primary_key)
                return float(rate)
            except Exception:
                rate, _ = self.data.get_rate(fallback_key)
                return float(rate)

        tech_ot_hr = _read_rate_with_fallback("tech. overtime", "tech. regular time")
        eng_ot_hr = _read_rate_with_fallback("eng. overtime", "eng. regular time")

        hours_per_day = 8
        tech_day_rate = tech_hr * hours_per_day
        eng_day_rate = eng_hr * hours_per_day

        machine_rows = []
        assignments: List[Assignment] = []
        eng_all: List[int] = []

        line_calc: Dict[str, Dict[str, int | bool]] = {}

        for s in selections:
            mi = self.data.models[s.model]
            base_training = ceil_int(s.qty / TRAINING_MACHINES_PER_DAY) if mi.training_applicable else 0
            training_days = base_training if s.training_required else 0

            tech_install_total = mi.tech_install_days_per_machine * s.qty
            tech_total = tech_install_total + training_days
            eng_training_potential = base_training if (mi.eng_days_per_machine > 0) else 0
            eng_training_days = eng_training_potential if s.training_required else 0
            eng_total = (mi.eng_days_per_machine * s.qty) + eng_training_days

            single_training = 1 if (s.training_required and mi.training_applicable) else 0
            if mi.tech_install_days_per_machine + single_training > window:
                raise ValueError(f"{s.model}: Install ({mi.tech_install_days_per_machine}) + Training ({single_training}) exceeds the Customer Install Window ({window}).")
            if mi.eng_days_per_machine > 0:
                single_eng_training = 1 if (s.training_required and mi.eng_days_per_machine > 0) else 0
                if mi.eng_days_per_machine + single_eng_training > window:
                    raise ValueError(
                        f"{s.model}: Engineer ({mi.eng_days_per_machine}) + Training ({single_eng_training}) exceeds the Customer Install Window ({window})."
                    )

            line_calc[s.model] = {
                "base_training": base_training,
                "training_days": training_days,
                "tech_total": tech_total,
                "eng_training_potential": eng_training_potential,
                "eng_training_days": eng_training_days,
                "eng_total": eng_total,
            }

            eng_headcount = 0

            machine_rows.append({
                "model": s.model,
                "qty": s.qty,
                "training_days": training_days,
                "training_potential": base_training,
                "training_required": s.training_required,
                "training_applicable": bool(mi.training_applicable),
                "eng_training_days": eng_training_days,
                "eng_training_potential": eng_training_potential,
                "tech_total": tech_total,
                "eng_total": eng_total,
                "tech_headcount": 0,
                "eng_headcount": eng_headcount,
            })

        # Engineer allocation: RPC models share a dedicated RPC engineer pool,
        # and never mix with non-RPC work.
        rpc_engineer_lines = [s for s in selections if s.model in RPC_MODELS and int(line_calc[s.model]["eng_total"]) > 0]
        non_rpc_engineer_lines = [s for s in selections if s.model not in RPC_MODELS and int(line_calc[s.model]["eng_total"]) > 0]

        rpc_eng_pool: List[int] = []
        rpc_eng_models: List[str] = []
        for s in rpc_engineer_lines:
            mi = self.data.models[s.model]
            info = line_calc[s.model]
            eng_alloc = chunk_allocate_by_machine(mi.eng_days_per_machine, s.qty, int(info["eng_training_days"]), window)
            rpc_eng_models.append(s.model)
            if not rpc_eng_pool:
                rpc_eng_pool = list(eng_alloc)
            else:
                needed = len(eng_alloc)
                if len(rpc_eng_pool) < needed:
                    rpc_eng_pool.extend([0] * (needed - len(rpc_eng_pool)))
                for i, d in enumerate(eng_alloc):
                    rpc_eng_pool[i] += d
            if rpc_eng_pool and max(rpc_eng_pool) > window:
                total_pool_days = sum(rpc_eng_pool)
                min_heads = ceil_int(total_pool_days / window)
                min_heads = max(min_heads, len(rpc_eng_pool))
                rpc_eng_pool = balanced_allocate(total_pool_days, min_heads)

        rpc_eng_pool = sorted(rpc_eng_pool, reverse=True)
        if rpc_eng_pool:
            rpc_model_set = sorted(set(rpc_eng_models))
            for idx, row in enumerate(machine_rows):
                if row["model"] in rpc_model_set:
                    machine_rows[idx]["eng_headcount"] = len(rpc_eng_pool)
            for i, d in enumerate(rpc_eng_pool, 1):
                assignments.append(Assignment(", ".join(rpc_model_set), "Engineer", i, d, d * eng_day_rate, "Engineer RPC"))
            eng_all.extend(rpc_eng_pool)

        for s in non_rpc_engineer_lines:
            info = line_calc[s.model]
            mi = self.data.models[s.model]
            eng_alloc = chunk_allocate_by_machine(mi.eng_days_per_machine, s.qty, int(info["eng_training_days"]), window)
            eng_all.extend(eng_alloc)
            for idx, row in enumerate(machine_rows):
                if row["model"] == s.model:
                    machine_rows[idx]["eng_headcount"] = len(eng_alloc)
            for i, d in enumerate(eng_alloc, 1):
                assignments.append(Assignment(s.model, "Engineer", i, d, d * eng_day_rate, "Engineer"))

        # Technician allocation with skills-matrix grouping.
        group_map = self._partition_tech_groups(selections)
        tech_group_loads: Dict[str, List[int]] = {}
        tech_group_members: Dict[str, List[str]] = {}

        for group_name, group_lines in group_map.items():
            if not group_lines:
                continue
            tech_group_members[group_name] = [x.model for x in group_lines]
            pool_loads: List[int] = []
            for s in group_lines:
                mi = self.data.models[s.model]
                info = line_calc[s.model]
                if int(info["tech_total"]) <= 0:
                    continue

                if self._is_generic_model_name(s.model):
                    total_days = int(info["tech_total"])
                    if pool_loads:
                        extra = balanced_allocate(total_days, len(pool_loads))
                        pool_loads = [pool_loads[i] + extra[i] for i in range(len(pool_loads))]
                    else:
                        pool_loads = balanced_allocate(total_days, 1)
                else:
                    alloc = chunk_allocate_by_machine(mi.tech_install_days_per_machine, s.qty, int(info["training_days"]), window)
                    if not pool_loads:
                        pool_loads = list(alloc)
                    else:
                        needed = len(alloc)
                        if len(pool_loads) < needed:
                            pool_loads.extend([0] * (needed - len(pool_loads)))
                        for i, d in enumerate(alloc):
                            pool_loads[i] += d
                    if pool_loads and max(pool_loads) > window:
                        total_pool_days = sum(pool_loads)
                        min_heads = ceil_int(total_pool_days / window)
                        min_heads = max(min_heads, len(pool_loads))
                        pool_loads = balanced_allocate(total_pool_days, min_heads)

            pool_loads = sorted(pool_loads, reverse=True)
            tech_group_loads[group_name] = pool_loads

            for idx, row in enumerate(machine_rows):
                if row["model"] in tech_group_members[group_name]:
                    machine_rows[idx]["tech_headcount"] = len(pool_loads) if int(row["tech_total"]) > 0 else 0

            for i, d in enumerate(pool_loads, 1):
                assignments.append(Assignment(
                    ", ".join(sorted(set(tech_group_members[group_name]))),
                    "Technician",
                    i,
                    d,
                    d * tech_day_rate,
                    group_name,
                ))

        tech_all = [d for loads in tech_group_loads.values() for d in loads]

        def _weekend_onsite_days(onsite_by_person: List[int], travel_in_day: int) -> int:
            weekend_days = 0
            for onsite_days in onsite_by_person:
                days = int(onsite_days or 0)
                if days <= 0:
                    continue
                onsite_start = int(travel_in_day) + 1
                onsite_end = onsite_start + days - 1
                for day in range(onsite_start, onsite_end + 1):
                    day_of_week = ((day - 1) % 7) + 1  # 1=Sun ... 7=Sat
                    if day_of_week in (1, 7):
                        weekend_days += 1
            return weekend_days

        tech_weekend_days = _weekend_onsite_days(tech_all, travel_in_day=1)
        eng_weekend_days = _weekend_onsite_days(eng_all, travel_in_day=eng_travel_in_day)

        tech_regular_days = max(0, int(sum(tech_all)) - tech_weekend_days)
        eng_regular_days = max(0, int(sum(eng_all)) - eng_weekend_days)
        tech_ot_hours = tech_weekend_days * hours_per_day
        eng_ot_hours = eng_weekend_days * hours_per_day
        tech_ot_day_rate = tech_ot_hr * hours_per_day
        eng_ot_day_rate = eng_ot_hr * hours_per_day

        tech_regular_cost = float(tech_regular_days) * tech_day_rate
        eng_regular_cost = float(eng_regular_days) * eng_day_rate
        tech_ot_cost = float(tech_ot_hours) * tech_ot_hr
        eng_ot_cost = float(eng_ot_hours) * eng_ot_hr

        tech = RoleTotals(len(tech_all), sum(tech_all), sorted(tech_all, reverse=True), tech_day_rate, tech_regular_cost + tech_ot_cost)
        eng = RoleTotals(len(eng_all), sum(eng_all), sorted(eng_all, reverse=True), eng_day_rate, eng_regular_cost + eng_ot_cost)

        trip_days_by_person = [a.onsite_days + TRAVEL_DAYS_PER_PERSON for a in assignments]
        n_people = len(trip_days_by_person)
        total_trip_days = sum(trip_days_by_person)
        total_hotel_nights = sum(max(d - 1, 0) for d in trip_days_by_person)

        exp_lines: List[ExpenseLine] = []

        def add_exp(name, qty, unit, detail):
            exp_lines.append(ExpenseLine(name, float(qty), float(unit), float(qty) * float(unit), detail))

        add_exp("Airfare", n_people, OVERRIDE_AIRFARE_PER_PERSON, f"{n_people} person(s) × {money(OVERRIDE_AIRFARE_PER_PERSON)}")
        add_exp("Baggage", total_trip_days, OVERRIDE_BAGGAGE_PER_DAY_PER_PERSON, f"{int(total_trip_days)} day(s) × {money(OVERRIDE_BAGGAGE_PER_DAY_PER_PERSON)}")

        parking, _ = self.data.get_rate("parking")
        car, _ = self.data.get_rate("car rental")
        hotel, _ = self.data.get_rate("hotel")
        per_diem, _ = self.data.get_rate("per diem weekday")
        prep, _ = self.data.get_rate("pre/post trip prep")
        travel_time_rate, _ = self.data.get_rate("travel time")

        add_exp("Car Rental", total_trip_days, car, f"{int(total_trip_days)} day(s) × {money(car)}")
        add_exp("Parking", total_trip_days, parking, f"{int(total_trip_days)} day(s) × {money(parking)}")
        add_exp("Hotel", total_hotel_nights, hotel, f"{int(total_hotel_nights)} night(s) × {money(hotel)}")
        add_exp("Per Diem", total_trip_days, per_diem, f"{int(total_trip_days)} day(s) × {money(per_diem)}")
        add_exp("Pre/Post Trip Prep", n_people, prep, f"{n_people} person(s) × {money(prep)}")
        travel_hours = 16 * n_people
        add_exp("Travel Time", travel_hours, travel_time_rate, f"{travel_hours} hr(s) × {money(travel_time_rate)}/hr")

        exp_total = sum(l.extended for l in exp_lines)
        max_onsite = max([a.onsite_days for a in assignments], default=0)
        grand_total = exp_total + tech.labor_cost + eng.labor_cost

        meta = {
            "machine_rows": machine_rows,
            "assignments": assignments,
            "window": window,
            "max_onsite": max_onsite,
            "n_people": n_people,
            "total_trip_days": total_trip_days,
            "exp_total": exp_total,
            "grand_total": grand_total,
            "skills_warning": self.skills_warning,
            "eng_travel_in_day": eng_travel_in_day,
            "tech_regular_days": tech_regular_days,
            "eng_regular_days": eng_regular_days,
            "tech_ot_hours": tech_ot_hours,
            "eng_ot_hours": eng_ot_hours,
            "tech_ot_rate": tech_ot_hr,
            "eng_ot_rate": eng_ot_hr,
            "tech_ot_days": tech_weekend_days,
            "eng_ot_days": eng_weekend_days,
            "tech_ot_day_rate": tech_ot_day_rate,
            "eng_ot_day_rate": eng_ot_day_rate,
            "tech_ot_cost": tech_ot_cost,
            "eng_ot_cost": eng_ot_cost,
        }
        return tech, eng, exp_lines, meta


    def _autosize_table_height(self, tbl, visible_rows=None, max_height=520):
        """Resize table height to fit contents (optionally cap by visible row count) to avoid inner scrolling."""
        try:
            tbl.resizeRowsToContents()
            header_h = tbl.horizontalHeader().height()
            frame = tbl.frameWidth() * 2
            total = header_h + frame + 12
            n = tbl.rowCount()
            if visible_rows is not None:
                n = min(n, int(visible_rows))
            for r in range(n):
                total += tbl.rowHeight(r)
            total = min(total, max_height)
            tbl.setMinimumHeight(total)
            tbl.setMaximumHeight(total)
        except Exception:
            pass


    
    
    def _clear_workload_calendar(self, rows: int = 6):
        rows = max(6, int(rows or 0))
        self.tbl_workload_calendar.clearContents()
        self.tbl_workload_calendar.setRowCount(rows)
        self.tbl_workload_calendar.setColumnCount(WORKLOAD_CALENDAR_DAYS)
        self.tbl_workload_calendar.setHorizontalHeaderLabels(
            [["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"][i % 7] for i in range(WORKLOAD_CALENDAR_DAYS)]
        )
        for r in range(rows):
            self.tbl_workload_calendar.setRowHeight(r, 26)
            for c in range(WORKLOAD_CALENDAR_DAYS):
                it = QTableWidgetItem("")
                it.setTextAlignment(Qt.AlignCenter)
                self.tbl_workload_calendar.setItem(r, c, it)

    def _render_workload_calendar(self, tech: RoleTotals, eng: RoleTotals, meta: Dict):
        """Render a 2-week Sun-Sat Gantt-style calendar (rows=people, cols=days)."""

        # Default assumptions:
        # - Tech travel-in: Sunday (day 1), first onsite Monday (day 2)
        # - Engineer travel-in depends on RPC mix (Sun/Monday/Tuesday rules)
        tech_travel_in = 1
        eng_travel_in = int(meta.get("eng_travel_in_day", 1) or 1)

        people = []
        for i, d in enumerate(tech.onsite_days_by_person, start=1):
            people.append((f"T{i}", int(d), tech_travel_in, QColor("#e04426")))
        for i, d in enumerate(eng.onsite_days_by_person, start=1):
            people.append((f"E{i}", int(d), eng_travel_in, QColor("#6790a0")))

        self._clear_workload_calendar(rows=len(people))
        labels = [p[0] for p in people] + [""] * max(0, self.tbl_workload_calendar.rowCount() - len(people))
        self.tbl_workload_calendar.setVerticalHeaderLabels(labels)

        for row, (_label, onsite_days, travel_in_day, base_color) in enumerate(people):
            if onsite_days <= 0:
                continue

            travel_color = QColor(base_color)
            travel_color.setAlpha(115)

            onsite_start = travel_in_day + 1
            onsite_end = onsite_start + onsite_days - 1
            travel_out = onsite_end + 1

            # Travel in/out
            for day in (travel_in_day, travel_out):
                if 1 <= day <= WORKLOAD_CALENDAR_DAYS:
                    item = self.tbl_workload_calendar.item(row, day - 1)
                    if item is None:
                        item = QTableWidgetItem("")
                        self.tbl_workload_calendar.setItem(row, day - 1, item)
                    item.setBackground(travel_color)

            # Onsite solid bar
            for day in range(onsite_start, onsite_end + 1):
                if 1 <= day <= WORKLOAD_CALENDAR_DAYS:
                    item = self.tbl_workload_calendar.item(row, day - 1)
                    if item is None:
                        item = QTableWidgetItem("")
                        self.tbl_workload_calendar.setItem(row, day - 1, item)
                    item.setBackground(base_color)

    def recalc(self):
        self._refresh_model_choices()
        if len(self.lines) == 0:
            self.reset_views()
            return
        try:
            tech, eng, exp_lines, meta = self.calc()
            if meta.get("skills_warning"):
                self.alert.setText(str(meta.get("skills_warning")))
                self.alert.show()
            else:
                self.alert.hide()

            self.card_tech.set_value(str(tech.headcount), f"{tech.total_onsite_days} total days")
            self.card_eng.set_value(str(eng.headcount), f"{eng.total_onsite_days} total days")
            self.card_window.set_value(f"{meta['max_onsite']} days", f"install window {meta['window']} days")
            self.card_total.set_value(money(meta["grand_total"]), "labor + expenses")

            self._render_workload_calendar(tech, eng, meta)
            self.lbl_total_val.setText(money(meta["grand_total"]))

            rows = meta["machine_rows"]
            self.tbl_breakdown.setRowCount(len(rows))
            for r_i, r in enumerate(rows):
                # Training display rules:
                # - If training is not applicable for this model, hide all training UI/labels.
                # - If applicable but user unchecked training, show “(training excluded)”.
                if not r.get("training_applicable", True):
                    tech_disp = str(r["tech_total"])
                else:
                    if r.get("training_required", True):
                        tech_disp = f"{r['tech_total']} (incl. {r['training_days']} Train)" if r.get("training_days", 0) > 0 else str(r["tech_total"])
                    else:
                        tech_disp = f"{r['tech_total']} (training excluded)"

                if r["eng_total"] == 0:
                    eng_disp = "—"
                elif not r.get("training_applicable", True):
                    eng_disp = str(r["eng_total"])
                else:
                    eng_tp = r.get("eng_training_potential", 0)
                    eng_td = r.get("eng_training_days", 0)
                    if r.get("training_required", True):
                        eng_disp = f"{r['eng_total']} (incl. {eng_td} Train)" if (eng_tp > 0 and eng_td > 0) else str(r["eng_total"])
                    else:
                        eng_disp = f"{r['eng_total']} (training excluded)" if eng_tp > 0 else str(r["eng_total"])

                vals = [r["model"], str(r["qty"]), tech_disp, eng_disp,
                        "—" if r["tech_headcount"] == 0 else str(r["tech_headcount"]),
                        "—" if r["eng_headcount"] == 0 else str(r["eng_headcount"])]
                for c, v in enumerate(vals):
                    it = QTableWidgetItem(v)
                    if c in [1, 4, 5]:
                        it.setTextAlignment(Qt.AlignCenter)
                    if c == 2 and r["training_required"]:
                        it.setForeground(Qt.darkYellow)
                    self.tbl_breakdown.setItem(r_i, c, it)

            assigns: List[Assignment] = meta["assignments"]
            self.tbl_assign.setRowCount(len(assigns))
            for i, a in enumerate(assigns):
                vals = [a.model, a.role, str(a.person_num), str(a.onsite_days), money(a.cost)]
                for c, v in enumerate(vals):
                    it = QTableWidgetItem(v)
                    if c in [2, 3]:
                        it.setTextAlignment(Qt.AlignCenter)
                    if c == 4:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.tbl_assign.setItem(i, c, it)

            self.tbl_labor.setRowCount(5)
            labor_rows = [
                ("Tech. Regular Time", money(tech.day_rate) + "/day", str(meta.get("tech_regular_days", tech.total_onsite_days)), str(tech.headcount), money((meta.get("tech_regular_days", tech.total_onsite_days)) * tech.day_rate)),
                ("Tech. Overtime (Sat/Sun)", money(meta.get("tech_ot_day_rate", 0.0)) + "/day", str(meta.get("tech_ot_days", 0)), str(tech.headcount), money(meta.get("tech_ot_cost", 0.0))),
                ("Eng. Regular Time", money(eng.day_rate) + "/day", str(meta.get("eng_regular_days", eng.total_onsite_days)), str(eng.headcount), money((meta.get("eng_regular_days", eng.total_onsite_days)) * eng.day_rate)),
                ("Eng. Overtime (Sat/Sun)", money(meta.get("eng_ot_day_rate", 0.0)) + "/day", str(meta.get("eng_ot_days", 0)), str(eng.headcount), money(meta.get("eng_ot_cost", 0.0))),
            ]
            for r_i, row in enumerate(labor_rows):
                for c, v in enumerate(row):
                    it = QTableWidgetItem(v)
                    if c in [2, 3]:
                        it.setTextAlignment(Qt.AlignCenter)
                    if c == 4:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.tbl_labor.setItem(r_i, c, it)


            # Subtotal row
            labor_subtotal = tech.labor_cost + eng.labor_cost
            self.tbl_labor.setItem(4, 0, QTableWidgetItem("Subtotal"))
            self.tbl_labor.setItem(4, 1, QTableWidgetItem(""))
            self.tbl_labor.setItem(4, 2, QTableWidgetItem(""))
            self.tbl_labor.setItem(4, 3, QTableWidgetItem(""))
            it = QTableWidgetItem(money(labor_subtotal))
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_labor.setItem(4, 4, it)

            self.lbl_exp_hdr.setText(
                f"Expenses are calculated using person-days, including {TRAVEL_DAYS_PER_PERSON} travel days per person."
            )
            self.tbl_exp.setRowCount(len(exp_lines) + 1)
            for i, l in enumerate(exp_lines):
                vals = [l.description, l.details, money(l.extended)]
                for c, v in enumerate(vals):
                    it = QTableWidgetItem(v)
                    if c == 2:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.tbl_exp.setItem(i, c, it)

            sub_row = len(exp_lines)
            self.tbl_exp.setItem(sub_row, 0, QTableWidgetItem("Expenses Subtotal"))
            self.tbl_exp.setItem(sub_row, 1, QTableWidgetItem("—"))
            it = QTableWidgetItem(money(meta["exp_total"]))
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_exp.setItem(sub_row, 2, it)

            self.btn_print.setEnabled(True)

        except Exception as e:
            self.reset_views()
            self.alert.setText(str(e))
            self.alert.show()

    def build_quote_html(self, tech: RoleTotals, eng: RoleTotals, exp_lines: List[ExpenseLine], meta: TDict[str, object]) -> str:
        from datetime import date, timedelta
        today = date.today()
        validity = today + timedelta(days=30)
        date_str = f"{today:%B} {today.day}, {today:%Y}"
        valid_str = f"{validity:%B} {validity.day}, {validity:%Y}"

        logo_html = ""
        if LOGO_PATH.exists():
            try:
                b = LOGO_PATH.read_bytes()
                b64 = base64.b64encode(b).decode("ascii")
                # Keep logo intentionally small for print layout; explicit attributes are
                # more reliable than class-only sizing in Qt rich-text rendering.
                logo_html = f'<img src="data:image/png;base64,{b64}" alt="Pearson" height="36" style="height:36px; width:auto; display:inline-block;" />'
            except Exception:
                logo_html = ""

        hdr_customer = str(self.quote_header_fields.get("customer_name", "")).strip()
        hdr_reference = str(self.quote_header_fields.get("reference", "")).strip()
        hdr_submitted = str(self.quote_header_fields.get("submitted_to", "")).strip()
        hdr_prepared = str(self.quote_header_fields.get("prepared_by", "")).strip()

        def _esc(x: str) -> str:
            return (
                x.replace("&", "&amp;")
                 .replace("<", "&lt;")
                 .replace(">", "&gt;")
                 .replace('"', "&quot;")
            )

        submitted_html = _esc(hdr_submitted).replace("\n", "<br/>") if hdr_submitted else ""

        mr = []
        for r in meta["machine_rows"]:
            if not r.get("training_applicable", True):
                tech_disp = str(r["tech_total"])
            else:
                if r.get("training_required", True):
                    tech_disp = f"{r['tech_total']} (incl. {r['training_days']} Train)" if r.get("training_days", 0) > 0 else str(r["tech_total"])
                else:
                    tech_disp = f"{r['tech_total']} (training excluded)"

            if r["eng_total"] == 0:
                eng_disp = "—"
            elif not r.get("training_applicable", True):
                eng_disp = str(r["eng_total"])
            else:
                eng_tp = r.get("eng_training_potential", 0)
                eng_td = r.get("eng_training_days", 0)
                if r.get("training_required", True):
                    eng_disp = f"{r['eng_total']} (incl. {eng_td} Train)" if (eng_tp > 0 and eng_td > 0) else str(r["eng_total"])
                else:
                    eng_disp = f"{r['eng_total']} (training excluded)" if eng_tp > 0 else str(r["eng_total"])

            mr.append(f"""<tr>
                <td>{r['model']}</td>
                <td style="text-align:center;">{r['qty']}</td>
                <td>{tech_disp}</td>
                <td style="text-align:center;">{eng_disp}</td>
                <td style="text-align:center;">{r['tech_headcount'] if r['tech_headcount'] else "—"}</td>
                <td style="text-align:center;">{r['eng_headcount'] if r['eng_headcount'] else "—"}</td>
            </tr>""")

        exp_rows = []
        for l in exp_lines:
            exp_rows.append(f"""<tr>
                <td>{l.description}</td>
                <td>{l.details}</td>
                <td style="text-align:right;">{money(l.extended)}</td>
            </tr>""")

        # Build printable workload calendar (same 14-day Gantt concept used in the UI).
        tech_travel_in = 1
        eng_travel_in = int(meta.get("eng_travel_in_day", 1) or 1)
        people = []
        for i, d in enumerate(tech.onsite_days_by_person, start=1):
            people.append((f"T{i}", int(d), tech_travel_in, "#e04426"))
        for i, d in enumerate(eng.onsite_days_by_person, start=1):
            people.append((f"E{i}", int(d), eng_travel_in, "#6790a0"))

        day_labels = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        cal_head = "".join([f"<th>{day_labels[i % 7]}</th>" for i in range(WORKLOAD_CALENDAR_DAYS)])

        def _lighten_hex(color_hex: str, amt: float = 0.58) -> str:
            color_hex = color_hex.strip().lstrip("#")
            if len(color_hex) != 6:
                return "#d7dce2"
            r = int(color_hex[0:2], 16)
            g = int(color_hex[2:4], 16)
            b = int(color_hex[4:6], 16)
            r = int(r + (255 - r) * amt)
            g = int(g + (255 - g) * amt)
            b = int(b + (255 - b) * amt)
            return f"#{r:02x}{g:02x}{b:02x}"

        cal_rows = []
        for label, onsite_days, travel_in_day, base_color in people:
            cells = ["<td></td>" for _ in range(WORKLOAD_CALENDAR_DAYS)]
            if onsite_days > 0:
                onsite_start = travel_in_day + 1
                onsite_end = onsite_start + onsite_days - 1
                travel_out = onsite_end + 1
                travel_color = _lighten_hex(base_color)
                for day in (travel_in_day, travel_out):
                    if 1 <= day <= WORKLOAD_CALENDAR_DAYS:
                        cells[day - 1] = f'<td class="cal-travel" style="background:{travel_color};"></td>'
                for day in range(onsite_start, onsite_end + 1):
                    if 1 <= day <= WORKLOAD_CALENDAR_DAYS:
                        cells[day - 1] = f'<td class="cal-onsite" style="background:{base_color};"></td>'
            cal_rows.append(f"<tr><th class=\"cal-person\">{label}</th>{''.join(cells)}</tr>")

        workload_calendar_html = f"""
            <h3>Workload Calendar</h3>
            <table class="grid grid-calendar">
                <tr><th class="cal-person">Person</th>{cal_head}</tr>
                {''.join(cal_rows) if cal_rows else f'<tr><td colspan="{WORKLOAD_CALENDAR_DAYS + 1}" class="muted">No personnel assigned.</td></tr>'}
            </table>
            <div class="muted">Legend: Light shade = travel day · Dark shade = onsite day</div>
        """

        labor_sub = tech.labor_cost + eng.labor_cost

        req_html = ""
        if self.data.requirements:
            li = "".join([f"<li>{x}</li>" for x in self.data.requirements])
            req_html = f"<h3>Requirements & Assumptions</h3><ul>{li}</ul>"

        header_html = f"""
            <table width="100%" role="presentation" style="border-collapse:collapse; margin:0 0 2px 0;">
                <tr>
                    <td></td>
                    <td align="right" valign="top" style="width:220px;">{logo_html}</td>
                </tr>
            </table>
            <table width="100%" class="topbar" role="presentation">
                <tr>
                    <td align="left" valign="top" style="padding-top:8px;">
                        <p class="title">Commissioning Budget Quote</p>
                        <p class="subtitle muted">Service Estimate</p>
                    </td>
                    <td align="right" valign="top" style="width:220px;"></td>
                </tr>
            </table>
        """

        html = f"""<html><head><meta charset="utf-8" />
        <style>
            @page {{ size: Letter; margin: 0.5in; }}
            body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10pt; color: #0F172A; }}
            .topbar {{ width: 100%; border-collapse: collapse; border-bottom: 3px solid #F05A28; margin: 0 0 12px 0; }}
            .topbar td {{ padding: 0 0 8px 0; }}
            .quote-logo {{ height: 36px; width: auto; display: inline-block; margin: 0; }}
            .title {{ font-size: 18pt; font-weight: 800; color: #4c4b4c; margin: 0; }}
            .subtitle {{ margin: 4px 0 0 0; color: #6D6E71; }}
            .grid {{ width: 100%; border-collapse: collapse; margin-top: 10px; table-layout: fixed; }}
            .grid th {{ background: #343551; color: white; text-align: left; padding: 8px; border-bottom: 1px solid #E2E8F0; }}
            .grid td {{ padding: 8px; border-bottom: 1px solid #E2E8F0; }}
            .grid-calendar th, .grid-calendar td {{ padding: 4px; text-align: center; }}
            .grid-calendar th.cal-person {{ width: 60px; text-align: left; }}
            .grid-calendar td.cal-travel {{ }}
            .grid-calendar td.cal-onsite {{ }}
            .box {{ border: 1px solid #E6E8EB; border-radius: 10px; padding: 10px; background: rgba(103,144,160,0.18); }}
            .summary-wrap {{ width: 100%; border-collapse: collapse; table-layout: fixed; border: 1px solid #E6E8EB; border-radius: 10px; background: rgba(103,144,160,0.18); }}
            .summary-wrap td {{ vertical-align: top; padding: 10px; }}
            .summary-wrap td.col-left {{ width: 48%; padding-right: 28px; }}
            .summary-wrap td.col-right {{ width: 52%; padding-left: 28px; }}
            .spacer-one-line {{ height: 12px; line-height: 12px; font-size: 1px; }}
            .new-page {{ page-break-before: always; }}
            h3 {{ color: #4c4b4c; margin: 18px 0 8px 0; }}
            .right {{ text-align: right; }}
            .muted {{ color: #6D6E71; }}
            .total {{ font-size: 16pt; font-weight: 900; color: #4c4b4c; }}
        </style></head><body>
            {header_html}

            <table class="summary-wrap" role="presentation">
                <tr>
                    <td class="col-left">
                        <b>Customer Name:</b><br/>{_esc(hdr_customer)}<br/><br/>
                        <b>Reference:</b><br/>{_esc(hdr_reference)}<br/><br/>
                        <b>Submitted to:</b><br/>{submitted_html}
                    </td>
                    <td class="col-right">
                        <b>Prepared By:</b><br/>{_esc(hdr_prepared)}<br/><br/>
                        <b>Quote Validity:</b><br/>{valid_str}<br/><br/>
                        <b>Total Personnel:</b><br/>{tech.headcount + eng.headcount} ({tech.headcount} Tech, {eng.headcount} Eng)<br/><br/>
                        <b>Estimated Duration:</b><br/>{meta["max_onsite"]} days onsite + {TRAVEL_DAYS_PER_PERSON} travel days
                    </td>
                </tr>
            </table>
            <div class="section-spacer"></div>

            <h3>Machine Breakdown</h3>
            <table class="grid">
                <tr><th>Model</th><th style="text-align:center;">Qty</th><th>Tech Days</th><th style="text-align:center;">Eng Days</th>
                    <th style="text-align:center;">Technicians</th><th style="text-align:center;">Engineers</th></tr>
                {''.join(mr)}
            </table>

            {workload_calendar_html}

            <div class="spacer-one-line">&nbsp;</div>
            <h3>Labor Costs</h3>
            <table class="grid">
                <tr><th>Item</th><th class="right">Extended</th></tr>
                <tr><td>Tech. Regular Time ({meta.get("tech_regular_days", tech.total_onsite_days)} days × {money(tech.day_rate)}/day)</td><td class="right">{money(meta.get("tech_regular_days", tech.total_onsite_days) * tech.day_rate)}</td></tr>
                <tr><td>Tech. Overtime (Sat/Sun) ({meta.get("tech_ot_days", 0)} day × {money(meta.get("tech_ot_day_rate", 0.0))}/day)</td><td class="right">{money(meta.get("tech_ot_cost", 0.0))}</td></tr>
                <tr><td>Eng. Regular Time ({meta.get("eng_regular_days", eng.total_onsite_days)} days × {money(eng.day_rate)}/day)</td><td class="right">{money(meta.get("eng_regular_days", eng.total_onsite_days) * eng.day_rate)}</td></tr>
                <tr><td>Eng. Overtime (Sat/Sun) ({meta.get("eng_ot_days", 0)} day × {money(meta.get("eng_ot_day_rate", 0.0))}/day)</td><td class="right">{money(meta.get("eng_ot_cost", 0.0))}</td></tr>
                <tr><td><b>Labor Subtotal</b></td><td class="right"><b>{money(labor_sub)}</b></td></tr>
            </table>

            <div class="new-page"></div>
            {header_html}
            <h3>Estimated Expenses</h3>
            <div class="muted">Includes {int(meta["total_trip_days"])} total trip day(s) across personnel (onsite + travel days).</div>
            <table class="grid">
                <tr><th>Expense</th><th>Details</th><th class="right">Amount</th></tr>
                {''.join(exp_rows)}
                <tr><td><b>Expenses Subtotal</b></td><td>—</td><td class="right"><b>{money(meta["exp_total"])}</b></td></tr>
            </table>

            <h3>Estimated Total</h3>
            <div class="box">
                <span class="total">{money(meta["grand_total"])}</span><br/>
                <span class="muted">Labor ({money(labor_sub)}) + Expenses ({money(meta["exp_total"])})</span>
            </div>

            <h3>Terms & Conditions</h3>
            <ul>
                <li><b>Pricing & Quote Expiration:</b> Prices shown reflect an estimate of days and expenses. Any additional time will be billed at the rates shown. Quote valid for 30 days.</li>
                <li><b>Customer Install Window:</b> No individual technician or engineer is assigned more than {meta["window"]} onsite days per trip.</li>
                <li><b>Training:</b> Training days are calculated at 1 day per {TRAINING_MACHINES_PER_DAY} machines of the same model type. Training can be excluded per machine if not required (customer request only).</li>
                <li><b>Machine-Specific Skills:</b> Each machine type requires technicians with specialized skills. Personnel are not shared across different machine types.</li>
                <li><b>Travel Days:</b> Expenses include {TRAVEL_DAYS_PER_PERSON} travel days (1 day travel-in + 1 day travel-out) in addition to onsite work days.</li>
            </ul>
            {req_html}
        </body></html>"""
        return html

    def print_quote_preview(self):
        try:
            tech, eng, exp_lines, meta = self.calc()
        except Exception as e:
            QMessageBox.critical(self, "Cannot print", str(e))
            return

        try:
            html = self.build_quote_html(tech, eng, exp_lines, meta)
            doc = QTextDocument()
            doc.setHtml(html)

            printer = QPrinter(QPrinter.HighResolution)
            # Letter page (8.5x11)
            try:
                printer.setPageSize(QPageSize(QPageSize.Letter))
            except Exception:
                pass

            preview = QPrintPreviewDialog(printer, self)
            preview.setWindowTitle("Print Preview - Commissioning Budget Quote")
            preview.setWindowModality(Qt.ApplicationModal)
            preview.resize(1100, 800)

            preview.paintRequested.connect(lambda p: doc.print_(p))

            # Show the dialog reliably
            preview.exec()
        except Exception as e:
            QMessageBox.critical(self, "Print error", str(e))
            return

    
    def _update_right_scroll_height_if_stacked(self):
        """When stacked, expand the right scroll area to its content so the OUTER scroll handles scrolling."""
        if not getattr(self, "_is_stacked", False):
            return
        try:
            if hasattr(self, "right_scroll") and hasattr(self, "right_content"):
                h = int(self.right_content.sizeHint().height()) + 80
                self.right_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                self.right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                self.right_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.right_scroll.setMinimumHeight(h)
                self.right_scroll.setMaximumHeight(h)
        except Exception:
            pass

    def _apply_responsive_layout(self):
        """Large screens: 2 columns w/ right-side scroll. Small screens: single stacked w/ full-window scroll."""
        if not hasattr(self, "splitter") or not hasattr(self, "outer_scroll") or not hasattr(self, "right_scroll"):
            return

        w = int(self.width())
        stacked = w < getattr(self, "_stack_threshold", 1280)

        if stacked and not getattr(self, "_is_stacked", False):
            self._is_stacked = True
            self.splitter.setOrientation(Qt.Vertical)
            # Enable whole-window scrolling; disable inner right scrolling
            self.outer_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.outer_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self._update_right_scroll_height_if_stacked()
            # Give left pane enough room; right pane will follow under it
            try:
                self.splitter.setSizes([650, 1000])
            except Exception:
                pass

            # In stacked (single-column) mode, make the machine configuration area taller so
            # multiple machine lines are visible without feeling cramped.
            try:
                if hasattr(self, "scroll"):
                    self.scroll.setMinimumHeight(240)
                    self.scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            except Exception:
                pass

        elif (not stacked) and getattr(self, "_is_stacked", False):
            self._is_stacked = False
            self.splitter.setOrientation(Qt.Horizontal)
            # Disable whole-window scrolling; allow right column to scroll
            self.outer_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.outer_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.right_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.right_scroll.setMinimumHeight(0)
            self.right_scroll.setMaximumHeight(16777215)
            self.right_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            try:
                self.splitter.setSizes([520, 1040])
            except Exception:
                pass

            # Restore default sizing for wide (two-column) mode.
            try:
                if hasattr(self, "scroll"):
                    self.scroll.setMinimumHeight(240)
            except Exception:
                pass

        elif stacked:
            # Still stacked; keep heights updated as content changes
            self._update_right_scroll_height_if_stacked()
            try:
                if hasattr(self, "scroll"):
                    self.scroll.setMinimumHeight(240)
            except Exception:
                pass

    def _apply_scale(self):
        # Scale UI typography modestly with window size; keep within sensible bounds.
        w = max(self.width(), 1)
        h = max(self.height(), 1)
        # Use width as primary driver; clamp to avoid extremes.
        scale = w / 1920.0
        scale = 0.85 if scale < 0.85 else (1.25 if scale > 1.25 else scale)
        pt = self._base_font_pt * scale
        f = QFont(self.font())
        f.setPointSizeF(pt)
        self.setFont(f)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._apply_responsive_layout()
        self._apply_scale()
    def closeEvent(self, event):

        event.accept()


def main():
    app = QApplication(sys.argv)
    w = MainWindow()
    w.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
